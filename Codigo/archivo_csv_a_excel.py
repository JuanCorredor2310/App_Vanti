from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, NamedStyle
global color_ele
import pandas as pd
color_ele = "56affe"

# * -------------------------------------------------------------------------------------------------------
# *                                             Almacenamiento excel básico
# * -------------------------------------------------------------------------------------------------------

def almacenar_csv_en_excel(df, archivo_xlsx, nombre_hoja):
    try:
        df.to_excel(archivo_xlsx, sheet_name=nombre_hoja, index=False)
        book = load_workbook(archivo_xlsx)
        sheet = book[nombre_hoja]
        generar_borde_simple(sheet)
        generar_borde_grueso(sheet)
        definir_tipo_texto(sheet)
        ajustar_ancho_columna(sheet)
        centrar_informacion(sheet)
        book.save(archivo_xlsx)
        return True
    except PermissionError:
        return None
    except ValueError:
        return None
    except Exception:
        return None
    except BaseException:
        return None

# * -------------------------------------------------------------------------------------------------------
# *                                             Almacenamiento excel con patrones
# * -------------------------------------------------------------------------------------------------------

def almacenar_csv_en_excel_patrones(df, archivo_xlsx, nombre_hoja):
    try:
        df.to_excel(archivo_xlsx, sheet_name=nombre_hoja, index=False)
        book = load_workbook(archivo_xlsx)
        sheet = book[nombre_hoja]
        generar_borde_simple(sheet)
        generar_borde_grueso(sheet)
        ajustar_ancho_columna(sheet)
        definir_tipo_texto(sheet)
        centrar_informacion(sheet)
        cambio_columnas(sheet)
        book.save(archivo_xlsx)
        return True
    except PermissionError:
        return None
    except ValueError:
        return None
    except Exception:
        return None
    except BaseException:
        return None

# * -------------------------------------------------------------------------------------------------------
# *                                             Personalización celdas excel
# * -------------------------------------------------------------------------------------------------------

def cambio_columnas(sheet):
    red_fill = PatternFill(start_color="FF5A5A", end_color="FF5A5A", fill_type="solid")
    for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for celda in fila:
            valor = celda.value
            valor = valor.replace("(","").replace(")","").replace("\'","").replace("np.int64","").replace("np.float64","")
            lista_valor = valor.split(",")
            if str(lista_valor[2].strip()) == "False":
                valor_1 = lista_valor[0]
                valor_2 = lista_valor[1]
                celda.fill = red_fill
                celda.value = str((valor_1,valor_2)).replace("\'","")
            else:
                celda.value = str(lista_valor[0].replace("\'",""))

def definir_tipo_texto(sheet):
    for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for celda in fila:
            valor = celda.value
            try:
                if isinstance(valor, int):
                    celda.value = int(str(int(celda.value)).replace("np.int64","").replace("np.float64",""))
                    celda.number_format = "0"
                elif isinstance(valor, float):
                    celda.value = float(str(float(celda.value)).replace("np.int64","").replace("np.float64",""))
                    celda.number_format = "0.00"
                else:
                    valor = str(valor).replace("np.int64","").replace("np.float64","")
                    celda.value = valor
                    celda.number_format = "General"
            except ValueError:
                valor = str(valor).replace("np.int64","").replace("np.float64","")
                celda.value = valor
                celda.number_format = "General"
            except BaseException:
                valor = str(valor).replace("np.int64","").replace("np.float64","")
                celda.value = valor
                celda.number_format = "General"

def ajustar_ancho_columna(sheet):
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter # Obtiene la letra de la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 10
        sheet.column_dimensions[column_name].width = adjusted_width

def centrar_informacion(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def generar_borde_grueso(sheet):
    fill = PatternFill(start_color=color_ele, end_color=color_ele, fill_type="solid")
    thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick'))
    for cell in sheet[1]:
        cell.fill = fill
        cell.border = thick_border

def generar_borde_simple(sheet):
    border_simple = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'))
    min_col = 1
    max_col = sheet.max_column
    min_row = 1
    max_row = sheet.max_row

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_simple

def lectura_hoja_xlsx(archivo, hoja):
    try:
        df = pd.read_excel(archivo, sheet_name=hoja).reset_index(drop=True)
        return df, True
    except FileNotFoundError:
        return None, None
    except PermissionError:
        return None, None
    except BaseException:
        return None, None

def hojas_disponibles(archivo):
    wb = load_workbook(archivo, read_only=True)
    lista_hojas = list(wb.sheetnames)
    return lista_hojas